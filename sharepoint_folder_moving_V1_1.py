from datetime import datetime, timedelta
from Nasiwak import*
import os
import shutil
import keyboard
import pandas as pd
import pyautogui
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from dateutil.relativedelta import relativedelta
from download_upload_delete import SharePointHandler
import tkinter as tk
import customtkinter as ctk
from tkcalendar import Calendar
from config_access_token import token_file


# from config import get_drive_id_from_site


CSVfolder = r'CSV'
excelfile = "access_data.xlsx"
excel_path = CSVfolder+'\\'+excelfile
builder_data="BuilderList.xlsx"


log_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(log_formatter)
logger = logging.getLogger()  # root logger
logger.setLevel(logging.INFO)
# Clear existing handlers to avoid duplicates
if logger.hasHandlers():
    logger.handlers.clear()
logger.addHandler(stream_handler)

# Replace with your actual file path
file_path = os.path.join(os.getcwd(), "Access_token", "Access_token.txt")
logging.info(f"file path for text file is: {file_path}")
# Open and read the file
with open(file_path, "r", encoding="utf-8") as file:
    content = file.read()
logging.info(f"Extracted text from .txt file is: {content}")

# Configuration
REPO_OWNER = "Nasiwak"
REPO_NAME = "sharepoint_transfer"
CURRENT_VERSION = "1.0.1"
ACCESS_TOKEN = content

Bot_Update(REPO_OWNER,REPO_NAME,CURRENT_VERSION,ACCESS_TOKEN)


# ACCESS_TOKEN = "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" # main token
webaccess_json_url = "https://raw.githubusercontent.com/Nasiwak/Nasiwak-jsons/refs/heads/main/webaccess.json"


webaccess_config = create_json_config(webaccess_json_url,ACCESS_TOKEN)

excelsheet = "結果.xlsx"

class data_fatch():
    def __init__(self,from_date,to_date) -> None:
        self.from_date = from_date
        self.to_date = to_date
        print(self.from_date,self.to_date)
        
        self.folder_with_date_time()
        self.create_styled_excel()
        self.create_clear_folder(CSVfolder)

        # create driver
        chrome_options = Options()
        chrome_options.add_experimental_option("prefs", {
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": False,
        })
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.maximize_window()

        # # Login web access
        self.wb=Webaccess(webaccess_config)
        self.wb.WebAccess_login(self.driver)

        # # download csv
        self.download_csv()
        
        # clean access data
        self.clean_access_data()
        # self.driver.close()
        
        self.process()

        self.driver.quit()



    def create_clear_folder(self,path):
        if os.path.exists(path):
            shutil.rmtree(path)
            logging.info("Folder removed.")
            os.makedirs(path)
            logging.info("Folder created.")
        else:
            os.makedirs(path)
            logging.info("Folder created.")

    def folder_with_date_time(self):
        cwd = os.getcwd()
        # Format: 案件_2025-04-17_1530
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        self.folder_name = f"案件_{timestamp}"
        full_path = os.path.join(cwd, self.folder_name)
        # logging.info(f"Folder created: {folder_name}")
        try:
            os.makedirs(full_path)
            logging.info(f"Folder created: {full_path}")
        except Exception as e:
            logging.info(f"Failed to create folder: {e}")

        return self.folder_name
       
       

        
    def clear_excel_data(self,Result):
            wb = load_workbook(Result)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=2):  # Assuming the first row has headers; adjust as needed
                    for cell in row:
                        cell.value = None  # Clear cell value
                        cell.fill = PatternFill(fill_type=None)  # Clear cell fill color
            wb.save(Result)
            logging.info("Data cleared from Excel file.")

    def csv_to_excel(self,csv_files):
        csv_files = [file for file in os.listdir(CSVfolder) if file.endswith(".csv")]
        try:
            # Create an empty DataFrame to hold all data
            all_data = pd.DataFrame()

            for csv_file in csv_files:
                csv_file_path = os.path.join(CSVfolder, csv_file)
                try:
                    # Read the CSV file
                    data = pd.read_csv(csv_file_path, encoding="CP932")
                    # Append the data to the all_data DataFrame
                    all_data = pd.concat([all_data, data], ignore_index=True)
                    logging.info(f"File {csv_file} converted and appended to the final sheet.")
                    os.remove(csv_file_path)
                except Exception as e:
                    logging.error(f"Failed to read and convert {csv_file}: {e}")
            
            # Save the combined data to an Excel file
            excel_file = os.path.join(CSVfolder, "access_data.xlsx")
            all_data.to_excel(excel_file, index=False, engine='openpyxl')
            logging.info(f"All CSV files combined and saved as {excel_file}")
            return excel_file
        except Exception as e:
            logging.info(f"Failed to create Excel file: {e}")

    def download_csv(self):
        受注一覧 = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧"]))).click()
        time.sleep(0.5)
        Reset = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["リセット"]))).click()
        time.sleep(5)
        for _ in range(5):  # Replace 5 with the number of times you want to press
            pyautogui.press('up')
        
        select_date1=WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["確定納品日_1"])))
        select_date1.clear()
        select_date1.send_keys(self.from_date)

        select_date=WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["確定納品日_2"])))
        select_date.clear()
        select_date.send_keys(self.to_date)
        # select_date.send_keys("2025/02/07")
        time.sleep(0.5)
        keyboard.press_and_release('tab')  # repeat as needed
        time.sleep(1)
        # select 新規
        dropdownbtn = self.driver.find_element(By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["出荷区分"])
        dropdownbtn.click()
            #checkbox path
        新規Checkbox =self.driver.find_element(By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["出荷区分_options"]["新規"])
        新規Checkbox.click()  
        time.sleep(0.5)
        keyboard.press_and_release('tab')
        time.sleep(0.5) 
       
        self.driver.execute_cdp_cmd('Page.setDownloadBehavior', {'behavior': 'allow', 'downloadPath': os.path.join(os.getcwd(),CSVfolder)})
        #search
        search_btn=self.driver.find_element(By.XPATH,webaccess_config["xpaths"]["受注一覧_xpaths"]["検索"])
        self.driver.execute_script("arguments[0].click();", search_btn)
        #Download file
        csv_download=WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["案件一覧のダウンロード"])))
        csv_download.click()
        time.sleep(4) 

        # run the code for covert Csv to excel
        self.exceldata=self.csv_to_excel(CSVfolder)

    def clean_access_data(self):
        builder_path = builder_data="BuilderList.xlsx"  # Assuming it's also in CSV folder

        # 🧾 Load Excel data
        access_df = pd.read_excel(excel_path)
        builder_df = pd.read_excel(builder_path)

        # 🏷️ Get the list of valid 得意先名 from builder file
        valid_names = builder_df['builder'].dropna().unique()

        # 🔍 Match and separate the data
        matched_df = access_df[access_df['得意先名'].isin(valid_names)]
        unmatched_df = access_df[~access_df['得意先名'].isin(valid_names)]

        # 🖨️ Print matched and unmatched values
        print("✅ 以下の得意先名はBuilderListに存在しているため残されました:")
        print(matched_df['得意先名'].unique())

        print("\n❌ 以下の得意先名はBuilderListに存在していないため削除されました:")
        print(unmatched_df['得意先名'].unique())

        # 💾 Save matched records back to the original file
        matched_df.to_excel(excel_path, index=False)

        logging.info(f"\nファイルが保存されました: {excel_path}")

    def fill_colour(self):
        # Define colors for highlighting
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # Keywords for red and green highlights
        red_keywords = ["NG", "納期無","ダウンロード無","error in anken"]
        green_keyword = "OK"
        # Iterate through all rows in the sheet
        for row in self.sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            for cell in row:
                if cell.value in red_keywords:
                    cell.fill = red_fill
                elif cell.value == green_keyword:
                    cell.fill = green_fill

        # Save the workbook
        # self.wb.save(os.path.join(excelsheet))
        # self.wb.close()
        
        
    def create_styled_excel(self):
        # 1. Generate timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        self.file_name = f"結果_{timestamp}.xlsx"
        file_path = os.path.join(os.getcwd(), self.file_name)

        # 2. Create a new Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 3. Header names
        headers = ["案件番号", "得意先名", "物件名", "確定納期", "", "結果","資料リンク"]

       # 4. Define Orange, Darker 50% fill (hex: #994D00)
        orange_fill = PatternFill(start_color="994D00", end_color="994D00", fill_type="solid")

        # # 5. Write headers and apply fill to all columns A to F
        # for col_num in range(1, 8):  # Columns 1 (A) to 6 (F)
        #     cell = ws.cell(row=1, column=col_num, value=headers[col_num - 1])
        #     cell.fill = orange_fill

        # # 6. Save the file
        # wb.save(file_path)
        # logging.info(f"Excel file created: {file_path}")
        # return self.file_name
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num, value=headers[col_num - 1])
            cell.fill = orange_fill

        # 6. Freeze the header row
        ws.freeze_panes = ws["A2"]

        # 7. Auto-adjust column widths based on header length
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # 8. Save the workbook
        wb.save(file_path)
        logging.info(f"Excel file created: {file_path}")
        return self.file_name
    
    def     update_new_sharepoint_link(self,bangou,link):
        try:
        
            受注一覧 = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧"]))).click()
            time.sleep(0.5)
            Reset = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["リセット"]))).click()
            time.sleep(5)

            search_bangou=WebDriverWait(self.driver,20).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["案件番号"])))
            logging.info("anken bangou box")
            search_bangou.send_keys(bangou)

            select_date1=WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, webaccess_config["xpaths"]["受注一覧_xpaths"]["確定納品日_1"])))
            select_date1.clear()
            time.sleep(0.5)
            keyboard.press_and_release('tab')
            time.sleep(0.5)
            search_btn=self.driver.find_element(By.XPATH,webaccess_config["xpaths"]["受注一覧_xpaths"]["検索"])
            self.driver.execute_script("arguments[0].click();", search_btn)

            # open anken 
            open=WebDriverWait(self.driver,20).until(EC.element_to_be_clickable((By.XPATH, "(//input[@type='button' and contains(@value, '参照')])[1]")))
            self.driver.execute_script("arguments[0].click();", open)
            
            logging.info("open")
            time.sleep(3)

            sharepoint_link_box = WebDriverWait(self.driver,20).until(EC.presence_of_element_located((By.XPATH, webaccess_config["xpaths"]["案件詳細_xpaths"]["Office365資料リンク"])))
            logging.info(f"locate sharepoint link box")
            sharepoint_link_box.clear()
            time.sleep(0.5)
            sharepoint_link_box.send_keys(link)
            time.sleep(1)
            logging.info(f"link updating")
            
            
            save=WebDriverWait(self.driver,20).until(EC.element_to_be_clickable((By.XPATH, webaccess_config['xpaths']["案件詳細_xpaths"]["案件情報を更新する"]))).click()
            logging.info("link updated")
           
            return True
        except Exception as e:
            logging.info(f"error while updating shiryou link : {e}")
            return False
            
        
    def process(self):
        df = pd.read_excel(excel_path, sheet_name='Sheet1')
        df = df.dropna(subset=['物件名'])
        excellinenumber = 2

        handler = SharePointHandler()  # Instantiate once before loop

        for _, row in df.iterrows():
            案件番号 = str(row['案件番号']) if not pd.isna(row['案件番号']) else row['案件番号']
            得意先名 = str(row['得意先名']) if not pd.isna(row['得意先名']) else row['得意先名']
            案件名 = str(row['物件名']) if not pd.isna(row['物件名']) else row['物件名']
            new_date = str(row['確定納期']) if not pd.isna(row['確定納期']) else row['確定納期']
            
            logging.info(f"案件番号: {案件番号}")
            time.sleep(3)
            
            
             # write reult in the excel sheet
            self.wb = load_workbook(os.path.join(self.file_name))
            self.sheet = self.wb["Sheet1"]
            self.sheet[f'A{excellinenumber}'].value = 案件番号
            self.sheet[f'B{excellinenumber}'].value = 得意先名
            self.sheet[f'C{excellinenumber}'].value = 案件名
            self.sheet[f'D{excellinenumber}'].value = new_date
            self.wb.save(self.file_name)
            time.sleep(1)
            # 2025/01/01-2025/01/06-2025/01/07-2025/01/08
            process, msg =handler.download_entire_folder(案件番号,self.folder_name)
            logging.info(f"{process} and {msg}")
            # process =handler.download_chizu_batch(案件番号)
            if process is True: 
                if msg == "none":
                    logging.info(f"新資料リンク無")
                    self.sheet[f'F{excellinenumber}'].value = "NG"
                    self.sheet[f'G{excellinenumber}'].value = msg
                else:
                    link_update=self.update_new_sharepoint_link(案件番号,msg)
                    if link_update is True:
                        self.sheet[f'F{excellinenumber}'].value = "OK"
                        self.sheet[f'G{excellinenumber}'].value = msg
                    else:
                        self.sheet[f'F{excellinenumber}'].value = "NG"
                        self.sheet[f'G{excellinenumber}'].value = msg
                    
                self.wb.save(self.file_name)
                time.sleep(1)
            else:
                self.sheet[f'F{excellinenumber}'].value = msg
                self.wb.save(self.file_name)
                time.sleep(1)
                
            excellinenumber += 1
            
            
class DateHandler:
    def __init__(self, from_date, to_date):
        self.from_date = from_date
        self.to_date = to_date
        self.process_data()

    def process_data(self):
        data_fatch(self.from_date,self.to_date)
        logging.info(f"From date: {self.from_date}")
        logging.info(f"To date: {self.to_date}")

        # Process the dates further as needed

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("SharePoint")
        self.root.geometry("400x350") 
        # self.root.configure(bg="lightgreen")
        ctk.set_appearance_mode("Light")
        # Create your widgets with the custom color theme
        self.heading_label = ctk.CTkLabel(self.root, text="SharePoint", font=("Arial", 24,"bold"), anchor="center", fg_color="#F5FFFA",bg_color="#ADD8E6" )
        self.heading_label.grid(row=0, column=0, columnspan=2, pady=20, sticky="nsew")

        self.from_date_label = ctk.CTkLabel(self.root, text="From Date : ", font=("Arial", 16), anchor="e",fg_color="#F5FFFA",bg_color="#ADD8E6")
        self.from_date_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")

        self.from_date_entry = ctk.CTkEntry(self.root, font=("Arial", 14))
        self.from_date_entry.grid(row=1, column=1, pady=5, sticky="nsw")

        # From Date Calendar below the entry box on the left side
        self.from_date_calendar = Calendar(self.root, selectmode='day', date_pattern='yyyy/mm/dd')
        self.from_date_calendar.grid(row=3, column=0, pady=5, sticky="ew")
        self.from_date_calendar.grid_forget()  # Hide the calendar initially

        # To Date Label and Entry
        self.to_date_label = ctk.CTkLabel(self.root, text="To Date : ", font=("Arial", 16), anchor="e",fg_color="#F5FFFA",bg_color="#ADD8E6")
        self.to_date_label.grid(row=3, column=0, padx=10, pady=5, sticky="e")

        self.to_date_entry = ctk.CTkEntry(self.root, font=("Arial", 14))
        self.to_date_entry.grid(row=3, column=1, pady=5, sticky="nsw")

        # To Date Calendar below the entry box
        self.to_date_calendar = Calendar(self.root, selectmode='day', date_pattern='yyyy/mm/dd')
        self.to_date_calendar.grid(row=4, column=0, columnspan=2, pady=5, sticky="ew")
        self.to_date_calendar.grid_forget()  # Hide the calendar initially

        # Start Button
        self.start_button = ctk.CTkButton(self.root, text="Start", command=self.on_start)
        self.start_button.grid(row=5, column=0, columnspan=2,padx=80, pady=20, sticky="nsew")

        # Version and Copyright Labels
        self.version_label = ctk.CTkLabel(self.root, text="Version 1.0", font=("Arial", 10), anchor="center")
        self.version_label.grid(row=6, column=0, columnspan=2, pady=5, sticky="nsew")

        self.copyright_label = ctk.CTkLabel(self.root, text="© 2025 Nasiwak", font=("Arial", 10), anchor="center")
        self.copyright_label.grid(row=7, column=0, columnspan=2, pady=5, sticky="nsew")

        # Bind the entry fields to show the calendar on click
        self.from_date_entry.bind("<Button-1>", self.show_from_date_calendar)
        self.to_date_entry.bind("<Button-1>", self.show_to_date_calendar)

        # Configure the rows and columns to expand evenly
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=1)
        self.root.grid_rowconfigure(3, weight=1)
        self.root.grid_rowconfigure(4, weight=1)
        self.root.grid_rowconfigure(5, weight=1)
        self.root.grid_rowconfigure(6, weight=1)
        self.root.grid_rowconfigure(7, weight=1)

        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

    def show_from_date_calendar(self, event):
        self.from_date_calendar.grid(row=2, column=0,columnspan=2, pady=5, sticky="ew")  # Place the calendar below the entry box
        self.to_date_calendar.grid_forget()  # Hide the To Date calendar if it's open
        # Bind calendar date selection to set date in entry and close the calendar
        self.from_date_calendar.bind("<<CalendarSelected>>", self.set_from_date)

    def show_to_date_calendar(self, event):
        self.to_date_calendar.grid(row=4, column=0, columnspan=2, pady=5, sticky="ew")  # Place the calendar below the entry box
        self.from_date_calendar.grid_forget()  # Hide the From Date calendar if it's open
        # Bind calendar date selection to set date in entry and close the calendar
        self.to_date_calendar.bind("<<CalendarSelected>>", self.set_to_date)

    def set_from_date(self, event):
        selected_date = self.from_date_calendar.get_date()
        self.from_date_entry.delete(0, tk.END)
        self.from_date_entry.insert(0, selected_date)
        self.from_date_calendar.grid_forget()  # Close the calendar after selection

    def set_to_date(self, event):
        selected_date = self.to_date_calendar.get_date()
        self.to_date_entry.delete(0, tk.END)
        self.to_date_entry.insert(0, selected_date)
        self.to_date_calendar.grid_forget()  # Close the calendar after selection

    def on_start(self):
        from_date = self.from_date_entry.get()
        to_date = self.to_date_entry.get()

        # Pass the selected dates to the class
        DateHandler(from_date, to_date)

if __name__ == "__main__":
    # Create Tkinter window and run the app
    root = ctk.CTk()
    app = App(root)
    root.mainloop()

        